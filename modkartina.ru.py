import os
import traceback
import asyncio
from collections import namedtuple

import openpyxl
import requests_html
from bs4 import BeautifulSoup


Option = namedtuple('Option', 'size price')
Category = namedtuple('Category', 'name parent_name url')
Product = namedtuple('Product', 'name category image_src code options')


async def get_categories() -> list[Category]:
    url = 'https://modkartina.ru/katalog-modulnyh-kartin'
    session = requests_html.AsyncHTMLSession()
    response = await session.get(url)
    html = response.html
    soup = BeautifulSoup(html.html, 'lxml')
    categories: list[Category] = []
    for category_tag in soup.find(id='column-left').li.ul.find_all('li'):
        parent_name = category_tag.a.text.strip()
        for child_tag in category_tag.find_all('li'):
            name = child_tag.text.strip('- \n')
            url = child_tag.find('a').get('href')
            category = Category(name, parent_name, url)
            categories.append(category)

    return categories


async def get_product(product_url, category):
    print(product_url)
    session = requests_html.AsyncHTMLSession()
    response: requests_html.HTMLResponse = await session.get(product_url)
    html = response.html
    soup = BeautifulSoup(html.html, 'lxml')
    name = soup.find(attrs={'itemprop': 'name'}).text.strip()
    code = soup.find(attrs={'itemprop': 'model'}).text.strip()
    script = '''
        () => {
            var prices = [];
            $("select").val('1').change();
            prices.push(price_product);
            $("select").val('2').change();
            prices.push(price_product);
            $("select").val('3').change();
            prices.push(price_product);
            $("select").val('4').change();
            prices.push(price_product);
            $("select").val('5').change();
            prices.push(price_product);
            return prices
        }
    '''
    try:
        await html.arender(script=script, timeout=100, keep_page=True)
    except Exception:
        await html.arender(script=script, timeout=100, keep_page=True)

    page = html.page
    prices = await page.evaluate(script)
    soup = BeautifulSoup(html.html, 'lxml')
    image_src = 'https://modkartina.ru/' + soup.find('image').get('xlink:href')
    options = []
    for price, option_tag in zip(prices, soup.find(id='_sV').find_all('option')):
        option = Option(option_tag.text.strip(), price)
        options.append(option)

    product = Product(name, category, image_src, code, options)
    await page.browser.close()
    return product


async def get_products(category: Category) -> list[Product]:
    session = requests_html.AsyncHTMLSession()
    response = await session.get(category.url)
    html = response.html
    soup = BeautifulSoup(html.html, 'lxml')
    products = []
    for product_tag in soup.find_all('div', class_='product-thumb'):
        try:
            product_url = product_tag.find(class_='caption').h4.a.get('href')
            if product_url == 'https://modkartina.ru/all-news/novost1':
                continue

            product = await get_product(product_url, category)
            products.append(product)
        except Exception:
            traceback.print_exc()
            break

    return products


def save_to_xlsx(products: list[Product], file_name):
    cur_dir = os.path.dirname(__file__)
    file_path = os.path.join(cur_dir, file_name)
    book = openpyxl.load_workbook(file_path)
    sheet = book['Лист1']
    for index, product in enumerate(products, 2):
        sheet.cell(index, 1).value = product.name
        sheet.cell(index, 2).value = product.code
        sheet.cell(index, 3).value = '/'.join([product.category.name,
                                               product.category.parent_name])
        sheet.cell(index, 4).value = product.image_src
        sheet.cell(index, 5).value = '\n'.join([' - '.join([str(option.size),
                                                            str(option.price)])
                                                for option in product.options])

    book.save(os.path.join(cur_dir, 'Результаты.xlsx'))


async def main():
    products = []
    for category in await get_categories():
        products += await get_products(category)
        break

    save_to_xlsx(products, 'Шаблон.xlsx')


if __name__ == "__main__":
    asyncio.run(main())
